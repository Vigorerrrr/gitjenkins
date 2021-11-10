# !/usr/bin python3                                 
# encoding   :utf-8 -*-                            
# @author    :贾克沣                              
# @software  :PyCharm      
# @file      :demo_excel_1.py
import ddt
"""
1,打开表单
2,读取标题数据
3,读取所有数据
4,指定单元格写入数据（使用静态方法，不要使用实例方法）


"""
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():

    def __init__(self,file):
        """  初始化参数  """
        self.file = file

    def open_sheet(self,name) -> Worksheet:
        """ 打开表单
        在函数/方法的后面加 -> 类型：表示此函数返回值是一个 这样的类型
        函数注解
        """
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self,sheet_name):
        """ 获取表单的表头 """
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read_all(self,sheet_name):
        """ 读取所有的数据 """
        sheet = self.open_sheet(sheet_name)
        data_rows = list(sheet.rows)[1:]
        # data_row = list(sheet.rows)[0]
        # headers = []
        # for i in data_row:
        #     headers.append(i.value)

        data = []
        for row in data_rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)

                data_dict = dict(zip(self.header(sheet_name),row_data))

            data.append(data_dict)
        return data

    @staticmethod
    def write(file,sheet_name,row,column,data_1):
        """ 写入excel数据 """
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row,column).value = data_1

        # 保存
        wb.save(file)

        wb.close()


if __name__ == '__main__':
    # excel = ExcelHandler(r"E:\123.xlsx")
    # data= excel.read_all('Sheet1')
    # print(data)
    # data = excel.write(r"E:\123.xlsx",'Sheet1',3,1,'new_value')
    pass





"""

ddt:数据驱动的思想，data driven testing
现在所说的是一个叫做 ddt 的 python 库
ddt 库是和unittest 搭配起来使用的，是unittest 的一个插件

@ddt.ddt
class TestDemo:

    @ddt.data(*test_data)
    将*test_data当中的一组测试数据，赋值到 data_info 这个参数
    def test_demo(self，data_info):
        pass
        
ddt解决了一个什么问题？
灵活的管理测试数据
用例的独立性和代码的灵活性 


"""



